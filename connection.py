import logging
import streamlit as st
import pandas as pd
from datetime import datetime, timedelta
from streamlit_drawable_canvas import st_canvas
from PIL import Image
import io
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.platypus import ListFlowable, ListItem
# from reportlab.platypus import (
#     SimpleDocTemplate, 
#     Table, 
#     TableStyle, 
#     Paragraph, 
#     Spacer, 
#     Image as ReportLabImage
# )
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as ReportLabImage
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.utils import ImageReader
from io import BytesIO
import os
import re
import traceback
import json
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseUpload
from datetime import date
import numpy as np


# 設置日誌記錄
logging.basicConfig(level=logging.INFO)

# 定義不同病房的藥品列表和庫存限制
WARD_DRUGS = {
    # "測試用選單": {
    #     "Morphine HCl 10mg/1mL/Amp":5,
    #     "Midazolam 15mg/3mL/Amp":5
    # },
    
    
    
    "ER": {
        "Morphine HCl 10mg/1mL/Amp": 20,
        "Lorazepam 2mg/mL/Amp": 5
    },
    "OR": {
        "Morphine HCl 10mg/1mL/Amp": 20
    },
    "麻醉科": {
        "MORPHINE20mg/1mL/Amp(PCA用)(5 amp/包)":10,
        "Morphine HCl 10mg/1mL/Amp":3,
        "Meperidine(Pethidine) 50mg/mL/Amp":3,
        "Fentanyl(0.05mg/mL) 2mL/Amp":400,
        "Fentanyl inj 0.05mg/mL 10mL/Amp":110,
        "Fentanyl inj 0.05mg/mL 10mL/Amp(PCA)(4 amp/包)":40,
        "Alfentanil 0.5mg/mL 2mL/Amp":100,
        #"Codeine phosphate 15mg/mL/Amp":,
        "Ketamine 500mg/10mL/Vial":4,
        #"Lorazepam 2mg/mL/Amp":,
        #"Midazolam 15mg/3mL/Amp":,
        "MIDazolam 五mg/mL/Amp":150,
        "Thiamylal 300mg/Amp":11,
        #"Diazepam 10mg/2mL/Amp":,
        "Propofol 200mg/20mL/Amp":600,
        "Etomidate 20mg/10mL/amp":15
    },
    
   "內視鏡": {
       "Fentanyl(0.05mg/mL) 2mL/Amp":40, #2026.01.22修改30->40
       "MIDazolam 五mg/mL/Amp":40        #2026.01.22修改30->40
    },
    
   "胸腔科檢查室": {
       "Meperidine(Pethidine) 50mg/mL/Amp":2,
       "MIDazolam 五mg/mL/Amp":10
    },
    
    "心導管": {
        "Morphine HCl 10mg/1mL/Amp":5,
        "Midazolam 15mg/3mL/Amp":5
    },
    
    "POR": {"Morphine HCl 10mg/1mL/Amp":10,
            "Meperidine(Pethidine) 50mg/mL/Amp":3,
            "Fentanyl(0.05mg/mL) 2mL/Amp":40,
            "Midazolam 15mg/3mL/Amp":2,
            "Diazepam 10mg/2mL/Amp":1
    },

    "6A": {
        "Morphine HCl 10mg/1mL/Amp": 20,
        "Meperidine(Pethidine) 50mg/mL/Amp": 3,
        "Codeine phosphate 15mg/imL INJ": 3,
        # "Lorazepam 2mg/mL/Amp": 35
    },
    
       
    "6A": {
        "Morphine HCl 10mg/1mL/Amp": 20,
        "Meperidine(Pethidine) 50mg/mL/Amp": 3,
        "Codeine phosphate 15mg/imL INJ": 3,
   
    },
    "MICU": {
        "Morphine HCl 10mg/1mL/Amp": 20,
        "Fentanyl inj 0.05mg/mL 10mL/Amp":60,
        "Lorazepam 2mg/mL/Amp": 2,
        "Propofol 200mg/20mL/Amp":20
        
    },
    "SICU": {
        "Morphine HCl 10mg/1mL/Amp": 20,
        "Meperidine(Pethidine) 50mg/mL/Amp": 5,
        "Fentanyl inj 0.05mg/mL 10mL/Amp":30,
        "Codeine phosphate 15mg/imL INJ": 5,
        "Lorazepam 2mg/mL/Amp": 2, 
        "Midazolam 15mg/3mL/Amp":2,
        "Propofol 200mg/20mL/Amp":10
    },
    "7A": {
        "Morphine HCl 10mg/1mL/Amp": 30,
        "Meperidine(Pethidine) 50mg/mL/Amp": 3,
        "Codeine phosphate 15mg/imL INJ": 5,

    },
    "7B": {
        "Morphine HCl 10mg/1mL/Amp": 20,
        "Meperidine(Pethidine) 50mg/mL/Amp": 3,

    },
    "8A": {
        "Morphine HCl 10mg/1mL/Amp": 30,
        "Meperidine(Pethidine) 50mg/mL/Amp": 2,

    },
    "8B": {
        "Morphine HCl 10mg/1mL/Amp": 20,


    },
    "9A": {
        "Morphine HCl 10mg/1mL/Amp": 20,
        "Meperidine(Pethidine) 50mg/mL/Amp": 3,
        "Codeine phosphate 15mg/imL INJ": 4,
        "Lorazepam 2mg/mL/Amp": 1
    },
    "9B": {
        "Morphine HCl 10mg/1mL/Amp": 15,

    }
       
}

# 定義列名
COLUMNS = ["現存量", "空瓶", "處方箋", "效期>6個月", "常備量=現存量+空瓶(空瓶量=處方箋量)", "備註"]

# 定義查核藥師列表
PHARMACISTS =[" ", "郭莉萱","洪英哲", "楊曜嘉", "廖文佑", "林昱男", "鍾向渝", "吳雨柔", "侯佳旻", "蘇宜萱", "簡妙格", "王奕祺", 
                  "陳意涵", "吳振凌", "周芷伊", "陳威如", "邱柏翰", "紀晨雲", "凃惠敏", "劉川葆", "江廷昌", "李典則", 
                  "盧柏融", "許家誠", "劉奕君", "張雯婷", "張亦汝", "林坤瑝", "洪繹婷", "呂奕芸", "黃穎慈", "張以靜", "陳薏帆", ] 

#口服管制藥品清單
oral_drugs = ["", "Flunitrazepam 2 mg/Tab", "Morphine Sulfate 15mg/Tab", "Codeine phosphate 30mg/T", 
                  "MORPHINE SULPHATE 30MG/T","Oxycodone HCL Immediate Release 5mg/Ca", "OxyContin Controlled-Release 20mg/Tab"]



# 設置 Google Drive API 認證
try:
    creds = service_account.Credentials.from_service_account_info(
        st.secrets["google_drive_credentials"],
        scopes=['https://www.googleapis.com/auth/drive.file']
    )
    drive_service = build('drive', 'v3', credentials=creds)
    st.write("Debug: Google Drive API 認證成功設置")
except Exception as e:
    st.error(f"設置 Google Drive API 認證時發生錯誤: {str(e)}")
    st.exception(e)

def upload_to_drive(file_name, mime_type, file_content):
    folder_id = st.secrets["google_drive"]["folder_id"]
    file_metadata = {
        'name': file_name,
        'parents': [folder_id]
    }
    media = MediaIoBaseUpload(file_content, mimetype=mime_type, resumable=True)
    file = drive_service.files().create(body=file_metadata, media_body=media, fields='id').execute()
    return file.get('id')

def create_drug_form(ward, drugs):
    data = {}
    incomplete_drugs = []
   
    for drug, limit in drugs.items():
        with st.expander(drug):
            drug_data = {}
            complete = True 
            reviewed = st.checkbox(f"✅ 已完成 {drug} 查核", key=f"{drug}_reviewed")
            for col in COLUMNS:
                if col == "現存量":
                    drug_data[col] = st.number_input(
                        f"{col} ({drug})",
                        min_value=0,
                        max_value=limit,
                        value=limit,
                        key=f"{drug}_{col}",
                        help=f"庫存限制: {limit}支"
                    )
                    if drug_data[col] > limit * 0.8:  # 如果庫存超過限制的80%，顯示警告
                        st.warning(f"注意：{drug}的庫存接近或超過限制（{limit}支）")
                        
                elif col in ["空瓶", "處方箋"]:
                    # 是否符合
                    status = st.radio(f"{col} 是否符合 ({drug})", ["符合", "不符合"], horizontal=True, key=f"{drug}_{col}_status")
                    if status == "符合":
                        # 自動計算 = 庫存上限 - 現存量
                        auto_value = max(limit - drug_data.get("現存量", 0), 0)
                        st.markdown(f"✅ 自動計算結果：**{auto_value}**")
                        drug_data[col] = auto_value
                    else:
                        # 讓使用者輸入數字
                        drug_data[col] = st.number_input(f"{col} ({drug})", min_value=0, value=0, key=f"{drug}_{col}_manual")


                
                elif col == "效期>6個月":
                    expiry_status = st.radio(f"{col} 是否符合 ({drug})", ["符合", "不符合"], horizontal=True, key=f"{drug}_{col}_status")
                    if expiry_status == "不符合":
                        expiry_reason = st.text_area(f"不符合原因 ({drug})", key=f"{drug}_{col}_reason")
                        drug_data[col] = f"不符合: {expiry_reason}" if expiry_reason else "不符合"
                        if not expiry_reason:
                            complete = False
                    else:
                        drug_data[col] = "符合"
                    st.markdown("---")

                
                elif col == "常備量=現存量+空瓶(空瓶量=處方箋量)":
                    stock_status = st.radio(f"{col} 是否符合 ({drug})", ["符合", "不符合"], horizontal=True, key=f"{drug}_{col}_status")
                    if stock_status == "不符合":
                        stock_reason = st.text_area(f"不符合原因 ({drug})", key=f"{drug}_{col}_reason")
                        drug_data[col] = f"不符合: {stock_reason}" if stock_reason else "不符合"
                        if not stock_reason:
                            complete = False
                        
                    else:
                        drug_data[col] = "符合"

                elif col == "備註":
                    drug_data[col] = st.text_area(f"{col} ({drug})", key=f"{drug}_{col}")
                    
                if col != "備註" and (drug_data[col] == "" or drug_data[col] is None):
                    complete = False


            # ✅【新增】檢查 checkbox 是否勾選
            if not reviewed:
                complete = False  # 未勾選則視為未完成

            if not complete:
                incomplete_drugs.append(drug)
                
            drug_data["已完成查核"] = reviewed

            data[drug] = drug_data                    


                

    return data, incomplete_drugs

def main():
# # ----------------------------------------------------
#     # ✅ 關鍵最終修正：嘗試使用 ListFlowable 及其子類別
#     try:
#         # 🚨 修正：直接從 reportlab.platypus 匯入 List 和 ListItem
#         from reportlab.platypus.flowables import ListFlowable as RLList 
#         from reportlab.platypus.listparagraphs import ListItem as RLListItem
#     except ImportError:
#         # 如果 ReportLab 列表模組匯入仍然失敗
#         st.error("🚨 錯誤：ReportLab 列表模組匯入失敗，請檢查環境設定。")
#         RLList = None
#         RLListItem = None
    
    
    if "oral_data_records" not in st.session_state:
        st.session_state.oral_data_records = []
    
    st.title("單位1-4級管制藥品月查核表")

    date_input_container = st.empty()

    # 獲取今天的日期
    today = date.today()

    # 使用唯一的 key 創建 date_input
    selected_date = date_input_container.date_input(
        "選擇日期",
        today,
        # max_value=today,
        key="date_input_unique_key"
    )

    # 選擇病房
    ward = st.selectbox("請選擇病房", list(WARD_DRUGS.keys()))

    # 獲取該病房的藥品列表和庫存限制
    drugs = WARD_DRUGS[ward]

    # 創建藥品表單
    data, incomplete_drugs = create_drug_form(ward, drugs)

    # ------------------------------------------------------------------------------------------------
    
    
    if "oral_data_records" not in st.session_state:
        st.session_state.oral_data_records = []
    
    
    # 2️⃣ 口服藥品查核 expander
    
    with st.expander(f"{ward} 口服管制藥品查核"):
        used_any = st.checkbox(f"單位是否有使用口服管制藥品", key=f"{ward}_used_any")
    
        if used_any:
            st.subheader("💊 新增口服藥品使用紀錄")
            
            # 1. 紀錄輸入欄位
            col1, col2 = st.columns(2)
            with col1:
                # 選擇查核藥品
                current_drug = st.selectbox("選擇查核藥品", oral_drugs, key=f"{ward}_select_drug_input")
                # 病人資訊
                current_bed = st.text_input(f"床號(填床號數字就好)", key=f"{ward}_oral_input_bed")
                current_mrn = st.text_input(f"病歷號", key=f"{ward}_oral_input_mrn")
                
            with col2:
                # 剩餘量查核
                current_expected = st.number_input(f"應剩餘量", min_value=0, value=0, step=1, key=f"{ward}_oral_input_expected")
                current_actual = st.number_input(f"實際剩餘量", min_value=0, value=0, step=1, key=f"{ward}_oral_input_actual")
                
                match = (current_expected == current_actual)
                current_reason = "" if match else st.text_area("不符合原因", key=f"{ward}_oral_input_reason")
    
            # 2. 添加按鈕邏輯
            if st.button(f"➕ 添加 {current_drug} 查核紀錄", key=f"{ward}_add_oral_record"):
                if not current_bed or not current_mrn:
                    st.warning("請輸入床號和病歷號。")
                else:
                    new_record = {
                        "查核藥品": current_drug,
                        "床號": current_bed,
                        "病歷號": current_mrn,
                        "應剩餘量": current_expected,
                        "實際剩餘量": current_actual,
                        "是否符合": "符合" if match else "不符合",
                        "不符合原因": current_reason,
                    }
                    st.session_state.oral_data_records.append(new_record)
                    st.success(f"已成功添加 {current_drug} / 床號 {current_bed} 的紀錄。")
                    # 重新運行以清空輸入欄位，準備下一筆資料
                    st.rerun()
            
            st.markdown("---")
            st.subheader("📝 已記錄的口服藥品查核清單")
            
            # 3. 顯示/刪除紀錄列表
            if st.session_state.oral_data_records:
                # 將列表轉換為 DataFrame 顯示，更清晰
                df_display = pd.DataFrame(st.session_state.oral_data_records)
                st.dataframe(df_display, use_container_width=True)
                
                # 提供刪除功能 (可選)
                if st.button("清空所有口服紀錄", key=f"{ward}_clear_oral"):
                    st.session_state.oral_data_records = []
                    st.rerun()
            else:
                st.info("目前沒有任何口服藥品使用紀錄。")
    
        else:
            st.info("本病房未使用口服管制藥品，可跳過查核")
            st.session_state.oral_data_records = [] # 如果取消勾選，清空紀錄


    # ------------------------------------------------------------------------------------------------
    
    # 添加查核藥師下拉選單
    pharmacist = st.selectbox("查核藥師", PHARMACISTS, help="請選擇查核藥師")

    # 添加電子簽名畫布
    st.write("請在下方簽名：")
    st.caption("使用滑鼠或觸控筆在下方空白處簽名")
    canvas_result = st_canvas(
        fill_color="rgba(255, 165, 0, 0.3)",
        stroke_width=2,
        stroke_color="#000000",
        background_color="#ffffff",
        height=300,
        drawing_mode="freedraw",
        key=f"canvas_{ward}",
    )
    # st.write(f"Debug: canvas_result.image_data: {canvas_result.image_data}") 檢查畫布內容用
    st.write("Debug: Starting main function")
    st.write(f"Debug: upload_to_drive function exists: {'upload_to_drive' in globals()}")
    
    # 在函數開始時初始化這些變量
    excel_filename = None
    pdf_filename = None
    excel_buffer = None
    pdf_buffer = None

    if st.button("提交", key="submit_button_unique_key"):
    # ✅【調試訊息】檢查 canvas 簽名 & 藥師選擇
        st.write(f"Debug: canvas_result.image_data is None: {canvas_result.image_data is None}")
        st.write(f"Debug: pharmacist: {pharmacist}")
        is_valid = True

        # ✅【1】檢查畫布是否有簽名
        if canvas_result.image_data is None or np.all(canvas_result.image_data == [255, 255, 255, 255]):  # 全白色表示沒有簽名
            st.error("❌ 請在畫布上簽名")
            is_valid = False

    
        # ✅【2】檢查是否選擇了藥師
        if pharmacist == "請選擇藥師" or pharmacist.strip() == "":
                st.error("❌ 請選擇查核藥師")
                is_valid = False

    
        # ✅【3】檢查是否有未填寫的藥品資料
        if incomplete_drugs:
                st.warning(f"🚨 以下藥品資料尚未填寫完整：{', '.join(incomplete_drugs)}")
                is_valid = False
# 🚀 只有所有檢查都通過 (is_valid == True)，才會執行下方的送出流程
        # ========================================================
        if is_valid:
            st.success(f"✅ 驗證成功！藥師 **{pharmacist}** 已完成提交。")

# ----------------------------------------------------
            # 🚀 數據準備區塊 (IV 藥品: df / 口服藥品: df_oral)
            # ----------------------------------------------------

            # 1. 取得日期和文件名
            file_date = selected_date.strftime("%Y.%m.%d")
            file_base_name = f"{file_date}_{ward}_單位1-4級管制藥品月查核表"
            excel_filename = f"{file_base_name}.xlsx"
            pdf_filename = f"{file_base_name}.pdf"
            
            # 2. 創建 IV 藥品 DataFrame (df) - 保持您的邏輯不變
            df_rows = []
            for drug, info in data.items():
                df_rows.append({
                    '單位': ward,
                    '常備品項': drug,
                    '常備量': WARD_DRUGS[ward][drug],
                    '現存量': info['現存量'],
                    '空瓶': info['空瓶'],
                    '處方箋': info['處方箋'],
                    '效期>6個月': info['效期>6個月'],
                    '常備量=現存量+空瓶(空瓶量=處方箋量)': info['常備量=現存量+空瓶(空瓶量=處方箋量)'],
                    '日期': selected_date.strftime("%Y/%m/%d"),
                    '被查核單位主管': '',
                    '查核藥師': pharmacist,
                    '備註': info['備註']
                })
            df = pd.DataFrame(df_rows)
            
            # 3. 創建 口服藥品 DataFrame (df_oral) - 修正後的邏輯
            
            # 讀取正確的紀錄列表
            oral_records = st.session_state.oral_data_records if 'oral_data_records' in st.session_state else []
            df_oral = pd.DataFrame() # 初始化 df_oral

            st.write("--- 口服資料調試 ---")
            st.write(f"紀錄數量: {len(oral_records)}")
            st.write("--------------------")

            if oral_records:
                # 確保 df_oral 使用正確的 List of Dicts
                df_oral = pd.DataFrame(oral_records)
                # 補上通用欄位
                df_oral.insert(0, '單位', ward) 
                df_oral['日期'] = selected_date.strftime("%Y/%m/%d")
                df_oral['查核藥師'] = pharmacist
            else:
                # 創建空的 df_oral，帶有所有欄位名稱
                df_oral = pd.DataFrame(columns=[
                    '單位', '查核藥品', '床號', '病歷號', 
                    '應剩餘量', '實際剩餘量', '查核結果', 
                    '不符合原因', '日期', '查核藥師'
                ])
                st.warning("⚠ 口服藥品沒有任何資料")

            
            # 保存為 Excel 文件
            excel_buffer = io.BytesIO()
            with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='單位1-4級管制藥品月查核表', index=False)
                
                # 調整列寬
                worksheet = writer.sheets['單位1-4級管制藥品月查核表']
                for idx, col in enumerate(df.columns):
                    max_length = max(df[col].astype(str).map(len).max(), len(col))
                    worksheet.column_dimensions[openpyxl.utils.get_column_letter(idx+1)].width = max_length + 2
                
                # ----------------------------------------------------
                # 寫入口服藥品到第二個 Sheet
                # ----------------------------------------------------
                if not df_oral.empty and len(df_oral) > 0:
                    sheet_name_oral = '口服查核資料'
                    df_oral.to_excel(writer, sheet_name=sheet_name_oral, index=False)
                    
                    # 調整 Oral Sheet 列寬
                    worksheet_oral = writer.sheets[sheet_name_oral]
                    for idx, col in enumerate(df_oral.columns):
                        # 使用 df_oral 的欄位來計算長度
                        max_length = max(df_oral[col].astype(str).map(len).max(), len(col))
                        worksheet_oral.column_dimensions[openpyxl.utils.get_column_letter(idx+1)].width = max_length + 2

                
                # 將簽名保存為圖片
                img = Image.fromarray(canvas_result.image_data.astype('uint8'), 'RGBA')
                img_byte_arr = io.BytesIO()
                img.save(img_byte_arr, format='PNG')
                img_byte_arr = img_byte_arr.getvalue()
                
                # 將簽名圖片添加到新的工作表
                worksheet = writer.book.create_sheet('被查核單位主管簽名')
                img = XLImage(io.BytesIO(img_byte_arr))
                worksheet.add_image(img, 'A1')
        
            # 生成 PDF 文件
            pdf_buffer = io.BytesIO()
            try:
                # 創建 PDF 文檔，使用 A4 橫向
                page_width, page_height = A4
                doc = SimpleDocTemplate(pdf_buffer, pagesize=(page_height, page_width), leftMargin=10*mm, rightMargin=10*mm, topMargin=10*mm, bottomMargin=10*mm)
                story = []
                styles = getSampleStyleSheet()
        
                # 註冊字體
                pdfmetrics.registerFont(TTFont('KaiU', 'fonts/kaiu.ttf'))  # 標楷體
                pdfmetrics.registerFont(TTFont('Calibri', 'fonts/calibri.ttf'))    # Calibri
        
                # 創建包含中文字體的樣式
                title_style = ParagraphStyle('TitleStyle', fontName='KaiU', fontSize=16, alignment=1)
                chinese_style = ParagraphStyle('ChineseStyle', fontName='KaiU', fontSize=9)
                english_style = ParagraphStyle('EnglishStyle', fontName='Calibri', fontSize=9)
                revision_style = ParagraphStyle('RevisionStyle', fontName='KaiU', fontSize=9, alignment=2)  # 改回右對齊
                small_title_style = ParagraphStyle('SmallTitle', fontName='KaiU', fontSize=7, leading=9, alignment=1)
                
                # 為 chinese_style 添加換行功能
                chinese_style.wordWrap = 'CJK'  # 支援中文自動換行
                chinese_style.leading = 10  # 設定行距
                
                # 為 english_style 添加換行功能
                english_style.wordWrap = 'CJK'  # 支援中文自動換行
                english_style.leading = 10  # 設定行距
                
                # 為 revision_style 添加換行功能
                revision_style.wordWrap = 'CJK'  # 支援中文自動換行
                revision_style.leading = 10  # 設定行距
                                
        
                # 添加查核時間、標題和修訂日期
                check_time = Paragraph("查核時間 : " + selected_date.strftime("%Y/%m/%d"), revision_style)  # 查核時間
                report_title = Paragraph("<b>單位1-4級管制藥品月查核表</b>", title_style)  # 標題，加粗處理
                update_time = Paragraph("更新時間 : 2026.4.20", revision_style)
        
                
                # 建立標題表格內容
                title_table_data = [
                    [check_time, report_title, update_time]  # 左 中 右 佈局
                ]
                                
                title_table = Table(title_table_data, colWidths=[page_height*0.2, page_height*0.6, page_height*0.2])
                title_table.setStyle(TableStyle([
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),  # 所有單元格居中對齊
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('LEFTPADDING', (0, 0), (-1, -1), 10),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 10),
                    ('TOPPADDING', (0, 0), (-1, -1), 0),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
                    ('SPAN', (1, 0), (1, 0)),  # 標題橫跨整行
                    ('ALIGN', (0, 1), (0, 1), 'LEFT'),  # 查核時間左對齊
                    ('ALIGN', (2, 1), (2, 1), 'RIGHT'),  # 修訂時間右對齊
                ]))
        
                story.append(title_table)
                story.append(Spacer(1, 5*mm))  # 減少標題和表格之間的間距
        
                # 創建簽名圖片
                img = ReportLabImage(BytesIO(img_byte_arr))
                img.drawHeight = 15*mm
                img.drawWidth = 30*mm
        
                # 準備表格數據
                # table_data = [
                #     ['單位', '常備品項', '常備量', '查核內容', '', '', '', '', '日期', '被查核單位主管', '查核藥師', '備註'],
                #     ['', '', '', '現存量', '空瓶', '處方箋', '常備量=現存量+空瓶(空瓶量=處方箋量)', '效期>6個月', '', '', '', '']
                # ]
        
                table_data = [
                            ['單位', '常備品項', '常備量', '查核內容', '', '', '', '', '日期', '被查核單位主管', '查核藥師', '備註'],
                            [
                                '', '', '', 
                                '現存量', '空瓶', '處方箋', 
                                Paragraph('常備量=現存量+空瓶(空瓶量=處方箋量)', small_title_style),  # 使用小字體樣式
                                Paragraph('效期>6個月', small_title_style),  # 讓「效期>6個月」標題也變小字體
                                '', '', '', ''
                            ]
                                ]        
        
                    
                                # 添加藥品數據
                for index, row in df.iterrows(): # <--- 修正 1: 使用 df 進行迭代
                    
                    # 🚨 修正 2: 根據 df 的欄位名來讀取資料
                    # 口服藥品的欄位可能為空 ('')，需要確保能處理
                    
                    # --- 處理 Paragraph 內容 (使用 row['欄位名稱'] 代替 info['...']) ---
                    # 這裡的邏輯需要注意，口服藥品的某些欄位（如常備量、空瓶、處方箋、效期判斷）可能是空的
                    
                    # 確保所有轉換都使用 row['欄位名稱']
                    expiry_paragraph = Paragraph(str(row['效期>6個月']), chinese_style)
                    stock_paragraph = Paragraph(str(row['常備量=現存量+空瓶(空瓶量=處方箋量)']), chinese_style)
                    remark_paragraph = Paragraph(str(row['備註']), chinese_style)
                    ward_paragraph = Paragraph(str(ward), chinese_style) # 單位可能還是使用 ward 變數
                
                    item_name = row['常備品項'] # <--- 從 df 中取出藥品名稱
                    
                    # --- 組裝 row 列表 ---
                    new_pdf_row = [
                        ward_paragraph, # 自動換行的「單位」
                        Paragraph(item_name, chinese_style), # 藥品名稱
                        
                        # 🚨 修正 3: 常備量需要從 df 讀取 (靜脈注射藥品有，口服藥品為空)
                        str(row['常備量']),
                        
                        # 🚨 修正 4: 其他欄位皆從 df 讀取
                        str(row['現存量']),
                        str(row['空瓶']),
                        str(row['處方箋']),
                        stock_paragraph, # 自動換行的「常備量=現存量+空瓶(空瓶量=處方箋量)」
                        expiry_paragraph, # 自動換行的「效期>6個月」
                        selected_date.strftime("%Y/%m/%d"),
                        img, # 確保 img 是在迴圈外定義且可用的
                        row['查核藥師'], # 從 df 讀取藥師名稱
                        remark_paragraph # 自動換行的「備註」
                    ]
                    
                    table_data.append(new_pdf_row) # <--- 將包含靜脈和口服藥品的行添加到表格數據中
        
                # 創建表格，調整列寬以適應 A4 橫向
                available_width = page_height - 10*mm
                col_widths = [10*mm, 45*mm, 10*mm, 10*mm, 10*mm, 10*mm, 49*mm, 40*mm, 20*mm, 30*mm, 20*mm, 23*mm]
                table = Table(table_data, colWidths=col_widths)
        
                # 設置表格樣式
                table.setStyle(TableStyle([
                    ('FONT', (0, 0), (-1, -1), 'KaiU'),
                    ('FONTSIZE', (0, 0), (-1, -1), 9),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('BACKGROUND', (0, 0), (-1, 1), colors.lightgrey),
                
                    # 合併「查核內容」標題
                    ('SPAN', (3, 0), (7, 0)),
                    ('SPAN', (9, 2), (9, -1)),
                    ('SPAN', (0, 2), (0, -1)),
                     # 合併「單位, 常備品項, 常備量」標題
                    ('SPAN', (0, 0), (0, 1)),
                    ('SPAN', (1, 0), (1, 1)),
                    ('SPAN', (2, 0), (2, 1)),
                    ('SPAN', (8, 0), (8, 1)),
                    ('SPAN', (9, 0), (9, 1)),
                    ('SPAN', (10, 0), (10, 1)),
                    ('SPAN', (11, 0), (11, 1)),
                    # 讓這些欄位內容自動換行
                    ('ALIGN', (6, 2), (6, -1), 'LEFT'),  # 效期>6個月
                    ('ALIGN', (7, 2), (7, -1), 'LEFT'),  # 常備量=現存量+空瓶(空瓶量=處方箋量)
                    ('ALIGN', (0, 2), (0, -1), 'LEFT'),  # 單位
                    ('ALIGN', (11, 2), (11, -1), 'LEFT'),  # 備註
                ]))
        

 # -----------------------------------------------------------------------------------------------以下為新增的口服PDF區塊

                story.append(table) # IV 藥品表格結束

                # ==============================================
                # 🚀 【新增】口服管制藥品使用查核區塊 (已清潔 U+00A0 字元)
                # ==============================================
                
                # 假設 oral_data 變數在上方已經定義並從 st.session_state 取得最新資料

                story.append(Spacer(1, 10*mm)) # 增加 IV 表格和新區塊的間距
                
                # 1. 設置口服藥品標題
                oral_title_style = ParagraphStyle('OralTitle', fontName='KaiU', fontSize=12, alignment=0, spaceAfter=5)
                story.append(Paragraph("<b>口服管制藥品使用查核</b>", oral_title_style))
                
                if not df_oral.empty: 
                    # 顯示「是」
                    oral_status_text = Paragraph("💊 本次查核口服管制藥品使用：**是**", chinese_style)
                    story.append(oral_status_text)
                    story.append(Spacer(1, 2*mm))

                    # 3. 遍歷 df_oral 的每一行並創建列表敘述
                    list_items = []
                    
                    # 🚨 關鍵修正：使用 df_oral.iterrows() 遍歷
                    for index, row in df_oral.iterrows():
                          # 1. 判斷是否「不符合」
                        is_not_match = row['是否符合'] == "不符合"
                        
                        # 2. 只有在不符合時才建立「不符合原因」的子句，否則為空字串
                        reason_clause = (
                            f" 不符合原因: {row['不符合原因']}. "
                            if is_not_match and row['不符合原因']
                            else ""
                        )            
                        # ✅ 關鍵修正：組合完整敘述，包含所有重要欄位
                        description = (
                            f"**{row['單位']}-{row['床號']} (病歷號: {row['病歷號']})** "
                            f"查核藥品: {row['查核藥品']}. "
                            f"結果: **{row['是否符合']}** (應剩餘量: {row['應剩餘量']}, 實際剩餘量: {row['實際剩餘量']})."
                            # ✅ 關鍵修正：將 reason_clause 條件式加入
                            f"{reason_clause}" 
                            f"日期: {row['日期']}. "
                            f"查核藥師: {row['查核藥師']}. "
                        )
                                            
                        list_items.append(
                            # ✅ 直接使用全局匯入的 ListItem
                            ListItem(Paragraph(description, chinese_style), leftIndent=20) 
                        )
                                        
                    # 4. 將列表加入 story (您可能漏掉了這一步驟)
                    story.append(
                        ListFlowable(
                            list_items, 
                            leftIndent=18,
                            # ✅ 關鍵修正：指定編號使用的字體為中文字體 (KaiU)
                            bulletFontName='KaiU', 
                            # 設置編號的大小 (通常與列表內容字體大小一致或略小)
                            bulletFontSize=10 
                        )
                    )
                    # if RLList is not None:
                    #     story.append(RLList(list_items, leftIndent=18)) # 使用 ReportLab 的 List 類
                    # else:
                    #     # 如果匯入失敗，我們至少可以在 PDF 中留下一個文字訊息
                    #     story.append(Paragraph("【錯誤：口服藥品列表模組載入失敗，無法生成列表。】", chinese_style))

                
                else:
                    # 顯示「否」
                    oral_status_text = Paragraph("💊 本次查核口服管制藥品使用：**否**", chinese_style)
                    story.append(oral_status_text)
                    story.append(Spacer(1, 2*mm))
                # ----------------------------------------------------
                # ==============================================
                # 口服藥品查核區塊結束
                # ==============================================


# ----------------------------------------------------------------------------------------------------以上為新增的口服PDF區塊
                # 生成 PDF
                doc.build(story)
                pdf_buffer.seek(0)
        
                st.write(f"Debug: excel_filename = {excel_filename}")
                st.write(f"Debug: pdf_filename = {pdf_filename}")
                st.write(f"Debug: excel_buffer is None: {excel_buffer is None}")
                st.write(f"Debug: pdf_buffer is None: {pdf_buffer is None}")
        
            except Exception as e:
                st.error(f"生成 PDF 時發生錯誤: {str(e)}")
                st.exception(e)
        
        # 在上傳文件之前檢查所有必要的變量是否已定義
        if excel_filename and pdf_filename and excel_buffer and pdf_buffer:
            st.write("Debug: 所有必要的變量都已設置")
            try:
                excel_file_id = upload_to_drive(excel_filename, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', excel_buffer)
                if excel_file_id:
                    st.success(f"Excel 文件已上傳，ID: {excel_file_id}")
                    excel_url = f"https://drive.google.com/file/d/{excel_file_id}/view"
                    st.markdown(f"[點擊此處查看 Excel 文件]({excel_url})")
                else:
                    st.error("Excel 文件上傳失敗")
        
                pdf_file_id = upload_to_drive(pdf_filename, 'application/pdf', pdf_buffer)
                if pdf_file_id:
                    st.success(f"PDF 文件已上傳，ID: {pdf_file_id}")
                    pdf_url = f"https://drive.google.com/file/d/{pdf_file_id}/view"
                    st.markdown(f"[點擊此處查看 PDF 文件]({pdf_url})")
                else:
                    st.error("PDF 文件上傳失敗")
            except Exception as e:
                st.error(f"上傳文件失敗: {str(e)}")
                st.exception(e)
        else:
            st.error("無法上傳文件：部分必要資訊缺失")
            st.write(f"Debug: excel_filename = {excel_filename}")
            st.write(f"Debug: pdf_filename = {pdf_filename}")
            st.write(f"Debug: excel_buffer is None: {excel_buffer is None}")
            st.write(f"Debug: pdf_buffer is None: {pdf_buffer is None}")

if __name__ == "__main__":
    main()






































































































